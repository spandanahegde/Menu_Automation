"""
run_analysis.py — orchestrates the raw-data-to-workbook pipeline:

  raw items + market intel
    -> role_classifier   (Weekday/Weekend Role + fit points)
    -> duplicate_detector (structural bundles, near-dup candidates, unique ingredients)
    -> analysis_engine    (Sections A-I -> every computed column)
    -> build_workbook_bytes() (65-column .xlsx, same schema report_engine.py reads)

RAW_COLUMNS is what the user's raw restaurant file must contain (Section 0
of the spec: Menu Item, Category, Ingredients, Total Qty Sold (annual),
Price, Theoretical Cost).

Confidence per stage — see the docstrings in role_classifier.py and
duplicate_detector.py for specifics. Short version: popularity, profit,
forecasting, and the recommendation rules themselves are validated exactly
against a real 54-item menu. Role assignment and duplicate detection are
heuristic and want a human (or an assistant) to sanity-check before you
trust the output completely — the app's review step exists for that.

What this does NOT generate: the creative REFRESH content (a new item
name + description + which 1-3 ingredients to add) — deciding what to
add to a dish is a judgment call, not a formula. This produces the
mechanical new PRICE for every REFRESH item, and leaves the creative
fields as an editable placeholder. The natural next step for those is to
hand the generated workbook to an assistant and ask it to draft the
refresh descriptions, the same way the original ones were written.
"""

import io
import re
from dataclasses import asdict

import openpyxl

import analysis_engine as ae
import role_classifier as rc
import duplicate_detector as dd
import menu_intelligence_ingest as mii
import sales_data_builder as sdb
import refresh_content_generator as rcg


RAW_COLUMNS = ['Menu Item', 'Category', 'Ingredients', 'Total Qty Sold (annual)',
               'Price', 'Theoretical Cost']

# Mirrors report_engine.REQUIRED_COLUMNS / the 65-column spec output.
OUTPUT_COLUMNS = [
    'No.', 'Current Menu Item', 'Current Category', 'Current Ingredients',
    'Current Price ($)', 'Theoretical Cost ($)', 'Gross Profit/Unit', 'Current Margin %',
    'Profitability Rank (1-5, relative)',
    'Future TC — 3 Mo ($)', 'Future TC — 6 Mo ($)', 'Future TC — 9 Mo ($)',
    'Future Margin — 3 Mo %', 'Future Margin — 6 Mo %', 'Future Margin — 9 Mo %',
    'Profitability Value — 3 Mo', 'Profitability Value — 6 Mo', 'Profitability Value — 9 Mo',
    'Commuter Score (0-100)', 'Commuter Score Percentile', 'Demographic Multiplier',
    'Category Modifier',
    'Popularity Rank — Absolute', 'Popularity Percentile (0-100)',
    'Current Profit — 3 Mo ($)', 'Current Profit — 6 Mo ($)',
    'Current Profit — 9 Mo ($)', 'Current Profit — 12 Mo ($)',
    'Baseline Growth Rate (%)', 'Market Multiplier', 'Adjusted Growth Rate (%)',
    'Forecast Qty — M1', 'Forecast Qty — M2', 'Forecast Qty — M3',
    'Forecast Qty — M4', 'Forecast Qty — M5', 'Forecast Qty — M6',
    'Forecast Sales — M1', 'Forecast Sales — M2', 'Forecast Sales — M3',
    'Forecast Sales — M4', 'Forecast Sales — M5', 'Forecast Sales — M6',
    'Total Forecast Qty — Next 6 Mo', 'Total Forecast Sales — Next 6 Mo ($)',
    'Current Qty — Last 6 Mo', 'Current Sales — Last 6 Mo ($)',
    'Qty Change — 6 Mo (Units)', 'Sales Change — 6 Mo ($)', 'Sales Change Signal',
    'Recommendation (KEEP/REFRESH/REMOVE)', 'Reason for Recommendation', 'Recommendation Explanation',
    'Comparable/Duplicate Menu Item Analysis', 'Ingredient Reuse Analysis',
    'Ingredients Reused from Current Menu', 'New Ingredients Required', 'Ingredients Removed',
    'Estimated Additional Food Cost ($)', 'New Theoretical Cost ($)', 'Suggested Selling Price ($)',
    'Suggested Refreshed Item', 'Suggested Refreshed Description',
    'Procurement Impact', 'Operational Complexity', 'Customer Approval Needed',
    'External Market Intelligence',
    'Weekday Role', 'Weekend Role',
    'Menu Performance Score (0-100)', 'Market Opportunity Score (0-100)',
]


def round_up_quarter(v):
    import math
    return math.ceil(v / 0.25) * 0.25


def compute_refresh_price(it):
    if it.profitability_rank <= 2:
        return max(it.theoretical_cost / (1 - 0.50), it.price)
    if it.profitability_rank == 3:
        return it.price + 1.50
    return max(it.price, it.price)  # Rank 4-5: no added-cost data -> no increase by default


def build_items_from_sales_data(file, market_override=None):
    """Reads a raw MICROS POS sales export (one row per item per month)
    and runs it through the same tested pipeline as build_items_from_raw,
    after first computing every Menu Intelligence financial field
    (Theoretical Cost resolution, Ingredient/Prep split, Profitability
    Value, Total Revenue/Profit, 3/6/9-month estimates) via
    sales_data_builder — see that module for the validated formulas.

    market_override: a MarketIntel to use (the sales export has no
    embedded market data, only a ZCTA — so this must come from somewhere:
    a manually-entered MarketIntel, a fresh Census/LODES fetch, or one
    pulled from a separate Menu Intelligence file that happens to have
    the same ZCTA).

    Returns (items, duplicate_candidates, financial_items, ingest_warnings).
    financial_items is the raw sales_data_builder output (has the
    Ingredient/Prep Cost breakdown etc., useful for a "Menu Intelligence"
    downloadable workbook even before the full recommendation engine runs).
    Raises ValueError if market_override isn't supplied.
    """
    financial_items, ingest_warnings = sdb.build_from_sales_csv(file)

    if market_override is None:
        raise ValueError(
            "The sales export has no embedded market data (only a ZCTA). "
            "Provide market data manually, via the Census/LODES fetch, or "
            "from a Menu Intelligence file for the same ZCTA."
        )

    raw_rows = [{
        'Menu Item': fi['name'], 'Category': fi['category'], 'Ingredients': '',
        'Total Qty Sold (annual)': fi['annual_qty'], 'Price': fi['price'],
        'Theoretical Cost': fi['theoretical_cost'],
    } for fi in financial_items]

    items, pairs = build_items_from_raw(raw_rows, market_override)
    return items, pairs, financial_items, ingest_warnings


def build_items_from_menu_intelligence(file, market_override=None):
    """Reads a 'Menu Intelligence' workbook (Menu Item/Category/Total Qty
    Sold/Avg Menu Price/Theoretical Cost or Ingredient+Prep Cost, usually
    with embedded ZCTA-level market data repeated per row) and runs it
    through the same tested pipeline as build_items_from_raw.

    market_override: a MarketIntel to use instead of whatever (if
    anything) was embedded in the file — e.g. if the user wants to
    override with a fresh Census/LODES fetch.

    Returns (items, duplicate_candidates, market_used, ingest_warnings).
    Raises ValueError if the file has no usable market data AND no
    market_override was supplied — the caller must get one from the user
    (manual entry or API fetch) in that case.
    """
    items_raw, market_from_file, ingest_warnings = mii.load_menu_intelligence_workbook(file)

    market = market_override or market_from_file
    if market is None:
        raise ValueError(
            "This file has no embedded market-intelligence data, and no "
            "market data was supplied another way. Provide it manually or "
            "via the Census/LODES fetch."
        )

    raw_rows = [{
        'Menu Item': r['name'], 'Category': r['category'], 'Ingredients': r['ingredients'],
        'Total Qty Sold (annual)': r['annual_qty'], 'Price': r['price'],
        'Theoretical Cost': r['theoretical_cost'],
    } for r in items_raw]

    items, pairs = build_items_from_raw(raw_rows, market)
    return items, pairs, market, ingest_warnings


def build_items_from_raw(raw_rows, market: ae.MarketIntel):
    """raw_rows: list of dicts with RAW_COLUMNS keys.
    Returns (items, duplicate_candidates) — items are fully-computed
    ItemInput objects; duplicate_candidates is the raw suggestion list for
    the review UI (before assign_duplicate_flags locks them in)."""
    items = []
    for idx, r in enumerate(raw_rows, start=1):
        it = ae.ItemInput(
            no=idx, name=str(r['Menu Item']).strip(), category=str(r['Category']).strip(),
            ingredients=str(r['Ingredients']).strip(),
            annual_qty=float(r['Total Qty Sold (annual)']), price=float(r['Price']),
            theoretical_cost=float(r['Theoretical Cost']),
            cost_estimated=bool(r.get('Theoretical Cost Estimated', False)),
        )
        ra = rc.assign_roles(it.category, it.price, it.name)
        it.weekday_role, it.weekday_desc = ra.weekday_role, ra.weekday_desc
        it.weekend_role, it.weekend_desc = ra.weekend_role, ra.weekend_desc
        it.weekday_fit_points, it.weekend_fit_points = ra.weekday_fit_points, ra.weekend_fit_points
        items.append(it)

    struct_bundles = dd.detect_structural_bundles(items)
    for it in items:
        it.is_structural_bundle = it.name in struct_bundles

    pairs = dd.detect_near_duplicates(items, threshold=0.5)
    qty_by_name = {it.name: it.annual_qty for it in items}
    dd.assign_duplicate_flags(items, pairs, qty_by_name)

    unique_flags = dd.detect_unique_ingredients(items)
    for it in items:
        it.unique_ingredients = unique_flags.get(it.name, False)

    ae.run_engine(items, market)

    for it in items:
        if it.recommendation == 'REFRESH':
            it.computed_new_price = round_up_quarter(compute_refresh_price(it))
        else:
            it.computed_new_price = None

    return items, pairs


def item_to_row(it, all_items=None, ingredient_index=None, use_llm=True, llm_warnings=None,
                 category_rotation=None):
    """all_items/ingredient_index: needed for the doc-5 formatters (full
    menu reference for duplicate/reuse checks) — build once per workbook
    via build_workbook_bytes, not per item. use_llm: attempt real
    Suggested Refreshed Item/Description/ingredient content via
    refresh_content_generator (needs ANTHROPIC_API_KEY); on any failure
    (no key, API error, validation failure) falls back to the existing
    '[NEEDS INPUT]' markers and appends a note to llm_warnings."""
    all_items = all_items if all_items is not None else [it]
    n_total = len(all_items)
    row = {
        'No.': it.no, 'Current Menu Item': it.name, 'Current Category': it.category,
        'Current Ingredients': it.ingredients,
        'Current Price ($)': it.price, 'Theoretical Cost ($)': it.theoretical_cost,
        'Gross Profit/Unit': it.gross_profit_per_unit, 'Current Margin %': it.current_margin_pct,
        'Profitability Rank (1-5, relative)': it.profitability_rank,
        'Future TC — 3 Mo ($)': it.future_tc_3m, 'Future TC — 6 Mo ($)': it.future_tc_6m,
        'Future TC — 9 Mo ($)': it.future_tc_9m,
        'Future Margin — 3 Mo %': it.future_margin_3m, 'Future Margin — 6 Mo %': it.future_margin_6m,
        'Future Margin — 9 Mo %': it.future_margin_9m,
        'Profitability Value — 3 Mo': it.profitability_value_3m,
        'Profitability Value — 6 Mo': it.profitability_value_6m,
        'Profitability Value — 9 Mo': it.profitability_value_9m,
        'Commuter Score (0-100)': it.commuter_score,
        'Commuter Score Percentile': it.commuter_score_percentile,
        'Demographic Multiplier': it.demographic_multiplier,
        'Category Modifier': it.category_modifier_value,
        'Popularity Rank — Absolute': it.popularity_rank,
        'Popularity Percentile (0-100)': it.popularity_percentile,
        'Current Profit — 3 Mo ($)': it.current_profit_3m, 'Current Profit — 6 Mo ($)': it.current_profit_6m,
        'Current Profit — 9 Mo ($)': it.current_profit_9m, 'Current Profit — 12 Mo ($)': it.current_profit_12m,
        'Baseline Growth Rate (%)': it.baseline_growth_rate * 100, 'Market Multiplier': it.market_multiplier,
        'Adjusted Growth Rate (%)': it.adjusted_growth_rate * 100,
        'Total Forecast Qty — Next 6 Mo': it.total_forecast_qty_6mo,
        'Total Forecast Sales — Next 6 Mo ($)': it.total_forecast_sales_6mo,
        'Current Qty — Last 6 Mo': it.current_qty_6mo, 'Current Sales — Last 6 Mo ($)': it.current_sales_6mo,
        'Qty Change — 6 Mo (Units)': it.qty_change_6mo, 'Sales Change — 6 Mo ($)': it.sales_change_6mo,
        'Sales Change Signal': it.sales_change_signal,
        'Recommendation (KEEP/REFRESH/REMOVE)': it.recommendation,
        'Weekday Role': f'{it.weekday_role} — {it.weekday_desc}',
        'Weekend Role': f'{it.weekend_role} — {it.weekend_desc}',
        'Menu Performance Score (0-100)': it.menu_performance_score,
        'Market Opportunity Score (0-100)': it.market_opportunity_score,
    }
    for i, (q, s) in enumerate(zip(it.forecast_qty_months, it.forecast_sales_months), start=1):
        row[f'Forecast Qty — M{i}'] = q
        row[f'Forecast Sales — M{i}'] = s

    signal_dict = {
        'pr': it.popularity_rank, 'pp': it.popularity_percentile,
        'cs': it.commuter_score_percentile, 'fr': it.profitability_rank,
        'p3': it.profitability_value_3m, 'p9': it.profitability_value_9m,
        'us': it.sales_change_signal,
        'sc': "{:,.0f}".format(row['Current Sales — Last 6 Mo ($)']),
        'sf': "{:,.0f}".format(row['Total Forecast Sales — Next 6 Mo ($)']),
        's': it.recommendation.lower(),
    }
    row['Reason for Recommendation'] = rcg.format_signal_reason(signal_dict)

    row['Comparable/Duplicate Menu Item Analysis'] = rcg.format_duplicate_analysis(
        it.name, it.duplicate_of, it.duplicate_outsells, n_total)
    if ingredient_index is not None:
        row['Ingredient Reuse Analysis'] = rcg.format_ingredient_reuse_analysis(
            it.name, it.ingredients, ingredient_index)
    else:
        row['Ingredient Reuse Analysis'] = ('Contains at least one menu-unique ingredient.'
                                             if it.unique_ingredients else 'No menu-unique ingredients.')

    row['Recommendation Explanation'] = row['Reason for Recommendation']

    if it.recommendation == 'REFRESH':
        creative = None
        if use_llm:
            try:
                creative = rcg.generate_refresh_creative_content(it, all_items, it.computed_new_price)
            except rcg.RefreshContentError as e:
                if llm_warnings is not None:
                    llm_warnings.append(f"'{it.name}': used the deterministic content "
                                         f"generator instead of the LLM ({e}).")
        if creative:
            # LLM path doesn't produce Ingredients Removed / cost breakdown
            # fields — fill those deterministically either way, since a
            # pure-addition refresh (no swap) is always a safe, valid answer.
            add1_cost = add2_cost = 0.15  # nominal if the LLM's own ingredient costs aren't known
            new_ing_count = 0 if str(creative['new_ingredients_required']).strip().lower().startswith('none') \
                else len([x for x in re.split(r',| and ', creative['new_ingredients_required']) if x.strip()])
            est_additional_cost = round(0.15 * max(new_ing_count, 1), 2)
            new_tc = round(it.theoretical_cost + est_additional_cost, 2)
            row['Suggested Refreshed Item'] = creative['suggested_refreshed_item']
            row['Suggested Refreshed Description'] = creative['suggested_refreshed_description']
            row['Ingredients Reused from Current Menu'] = creative['ingredients_reused']
            row['New Ingredients Required'] = creative['new_ingredients_required']
            row['Ingredients Removed'] = 'None — pure addition, original recipe fully retained.'
            row['Estimated Additional Food Cost ($)'] = est_additional_cost
            row['New Theoretical Cost ($)'] = new_tc
            row['Suggested Selling Price ($)'] = it.computed_new_price
            row['Procurement Impact'] = rcg.compute_procurement_impact('refresh', new_ingredient_count=new_ing_count)
            row['Operational Complexity'] = (
                'Low — 0-1 new ingredients' if new_ing_count <= 1 else
                f'Moderate — {new_ing_count} new ingredients, confirm prep flow before menu print')
        else:
            rotation_index = category_rotation.get(it.category, 0) if category_rotation is not None else 0
            fallback = rcg.generate_deterministic_refresh_content(it, rotation_index=rotation_index)
            if category_rotation is not None:
                category_rotation[it.category] = rotation_index + 2
            row['Suggested Refreshed Item'] = fallback['suggested_refreshed_item']
            row['Suggested Refreshed Description'] = fallback['suggested_refreshed_description']
            row['Ingredients Reused from Current Menu'] = fallback['ingredients_reused']
            row['New Ingredients Required'] = fallback['new_ingredients_required']
            row['Ingredients Removed'] = fallback['ingredients_removed']
            row['Estimated Additional Food Cost ($)'] = fallback['estimated_additional_food_cost']
            row['New Theoretical Cost ($)'] = fallback['new_theoretical_cost']
            row['Suggested Selling Price ($)'] = fallback['suggested_selling_price']
            row['Procurement Impact'] = fallback['procurement_impact']
            row['Operational Complexity'] = fallback['operational_complexity']
        row['Customer Approval Needed'] = 'Yes — new recipe/price'
    elif it.recommendation == 'KEEP':
        row['Suggested Refreshed Item'] = 'N/A — Item Retained'
        row['Suggested Refreshed Description'] = 'Item retained as-is — no refresh needed.'
        row['Ingredients Reused from Current Menu'] = 'N/A'
        row['New Ingredients Required'] = 'N/A'
        row['Ingredients Removed'] = 'N/A'
        row['Estimated Additional Food Cost ($)'] = 0.0
        row['New Theoretical Cost ($)'] = it.theoretical_cost
        row['Suggested Selling Price ($)'] = it.price
        row['Procurement Impact'] = rcg.compute_procurement_impact('keep')
        row['Operational Complexity'] = 'None — retained as-is'
        row['Customer Approval Needed'] = 'No'
    else:
        removed_ing_count = len(rcg._parse_ingredients(it.ingredients))
        row['Suggested Refreshed Item'] = 'N/A — Item Removed'
        row['Suggested Refreshed Description'] = f'Removed from menu. {row["Reason for Recommendation"]}'
        row['Ingredients Reused from Current Menu'] = 'N/A'
        row['New Ingredients Required'] = 'N/A'
        row['Ingredients Removed'] = 'N/A — item discontinued, not modified.'
        row['Estimated Additional Food Cost ($)'] = 0.0
        row['New Theoretical Cost ($)'] = 0.0
        row['Suggested Selling Price ($)'] = 0.0
        row['Procurement Impact'] = rcg.compute_procurement_impact('remove', removed_ingredient_count=removed_ing_count)
        row['Operational Complexity'] = 'None — removed from menu'
        row['Customer Approval Needed'] = 'No'

    row['External Market Intelligence'] = (
        f"Weekday: {it.weekday_role} — {it.weekday_desc} | Weekend: {it.weekend_role} — {it.weekend_desc}")

    return row


MENU_INTELLIGENCE_COLUMNS = [
    'Menu Item', 'Category', 'Total Qty Sold', 'Avg Menu Price ($)',
    'Theoretical Cost ($)', 'Ingredient Cost ($)', 'Prep Cost ($)',
    'Profitability Value ($)', 'Profitability Value (%)',
    'Total Revenue ($)', 'Total Profit ($)',
    'Est. Profit – 3M ($)', 'Est. Profit – 6M ($)', 'Est. Profit – 9M ($)',
    'ZCTA', 'TC Source',
]


def build_menu_intelligence_workbook_bytes(financial_items, market=None):
    """Builds a standalone 'Menu Intelligence' workbook (matching the
    reference schema's columns) from sales_data_builder's output — an
    intermediate, inspectable artifact before the full recommendation
    engine runs. If a MarketIntel is supplied, its fields are appended as
    repeated columns on every row (matching how the reference file embeds
    ZCTA-level market data per-row)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Menu Intelligence'

    market_cols = []
    if market is not None:
        market_cols = [
            'Residents (Living In Area)', 'Employed In Area', 'Worker Inflow',
            'Resident Outflow', 'Stay Local', 'Inflow Earning Over 3333 Pct',
            'Inflow Earning Up To 1250 Pct', 'Inflow Age 30 To 54 Pct', 'Inflow Age 55 Plus Pct',
        ]
    all_cols = MENU_INTELLIGENCE_COLUMNS + market_cols

    for c, col in enumerate(all_cols, start=1):
        ws.cell(1, c, col)  # header on row 1 — no leading blank rows
    for r, fi in enumerate(financial_items, start=2):
        row = {
            'Menu Item': fi['name'], 'Category': fi['category'], 'Total Qty Sold': fi['annual_qty'],
            'Avg Menu Price ($)': fi['price'], 'Theoretical Cost ($)': fi['theoretical_cost'],
            'Ingredient Cost ($)': fi['ingredient_cost'], 'Prep Cost ($)': fi['prep_cost'],
            'Profitability Value ($)': fi['profit_value'], 'Profitability Value (%)': fi['profit_pct'],
            'Total Revenue ($)': fi['total_revenue'], 'Total Profit ($)': fi['total_profit'],
            'Est. Profit – 3M ($)': fi['est_profit_3m'], 'Est. Profit – 6M ($)': fi['est_profit_6m'],
            'Est. Profit – 9M ($)': fi['est_profit_9m'], 'ZCTA': fi['zcta'], 'TC Source': fi['tc_source'],
        }
        if market is not None:
            row.update({
                'Residents (Living In Area)': market.residents, 'Employed In Area': market.daytime_workers,
                'Worker Inflow': market.worker_inflow, 'Resident Outflow': market.resident_outflow,
                'Stay Local': market.stay_local, 'Inflow Earning Over 3333 Pct': market.pct_income_high / 100,
                'Inflow Earning Up To 1250 Pct': market.pct_income_low / 100,
                'Inflow Age 30 To 54 Pct': market.pct_age_mid / 100,
                'Inflow Age 55 Plus Pct': market.pct_age_senior / 100,
            })
        for c, col in enumerate(all_cols, start=1):
            ws.cell(r, c, row.get(col, ''))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_workbook_bytes(items, use_llm=True, llm_warnings=None):
    """Builds an .xlsx matching the schema report_engine.py's load_rows()
    expects (header row with 'No.' in col A), returns bytes ready to feed
    straight into report_engine.generate_report(). use_llm/llm_warnings:
    see item_to_row — set use_llm=False to skip API calls entirely (e.g.
    no ANTHROPIC_API_KEY configured) and go straight to the always-
    available deterministic content generator for every REFRESH row (see
    refresh_content_generator.generate_deterministic_refresh_content) —
    every REFRESH row is guaranteed complete either way, never a
    '[NEEDS INPUT]' placeholder."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Menu Refresh Analysis'
    for c, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(1, c, col)  # header on row 1 — no leading blank rows
    ingredient_index = rcg.build_ingredient_index(items)
    category_rotation = {}
    for r, it in enumerate(items, start=2):
        row = item_to_row(it, all_items=items, ingredient_index=ingredient_index,
                           use_llm=use_llm, llm_warnings=llm_warnings,
                           category_rotation=category_rotation)
        for c, col in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(r, c, row.get(col, ''))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf