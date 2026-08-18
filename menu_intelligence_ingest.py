"""
menu_intelligence_ingest.py — reads the "Menu Intelligence" workbook schema
(Menu Item / Category / Total Qty Sold / Popularity Percentile / Avg Menu
Price / Theoretical Cost / Ingredient Cost / Prep Cost / ... / ZCTA +
per-row commuter-flow columns) and turns it into ItemInput objects plus a
MarketIntel, ready for analysis_engine.run_engine().

Two things this solves:

1. "Theoretical values generated using formulas like ingredient + prep
   cost." Theoretical Cost is resolved per row with this priority:
     a. Use Theoretical Cost as given, if present and plausible (a positive
        number, and food-cost% = TC/Price falls in a sane 5-70% band).
     b. Else, Theoretical Cost = Ingredient Cost + Prep Cost, if both are
        present.
     c. Else, Theoretical Cost = Price * category benchmark food-cost%
        (CATEGORY_FC_BENCHMARK below — reused from sales_data_builder.py,
        the same table input_normalizer.py and menu_estimation.py already
        use, so this module's fallback estimates match every other
        ingestion path instead of drifting from them. Replace
        sdb.CATEGORY_FC_BENCHMARK in one place for a different
        restaurant's cost structure.)
     d. Else (no category benchmark either), Theoretical Cost = Price * 25%
        as a last-resort flat default.
   Every row's resolution method is recorded (row['tc_source']) so you can
   see which values were given vs. estimated.

2. "Even when some columns from my code aren't in the uploaded file, it
   still needs to generate a workbook." Column names are matched via
   COLUMN_ALIASES (case-insensitive, tolerant of the naming variants seen
   across the different files in this project — 'Current Menu Item' vs
   'Menu Item', 'Total Qty Sold (annual)' vs 'Total Qty Sold', etc.).
   Anything genuinely absent (there's no Ingredients column in this
   schema) is defaulted rather than raising — Ingredients defaults to ''
   (duplicate/unique-ingredient detection just runs at lower fidelity,
   using name+category only, and says so).

Market intelligence: this schema repeats the same ZCTA-level commuter-flow
figures on every row (Employed In Area, Living In Area, Inflow/Outflow/
Interior breakdowns by age/earnings/industry). That's read ONCE from the
first valid row and turned into a MarketIntel — no API call needed when
this data is already embedded in the file. Verified against the project's
known reference numbers for ZCTA 38103 (residents 6,462; daytime workers
45,092; worker inflow 43,641; resident outflow 5,011; stay-local 1,451;
66.15% high income / 12.95% low income / 57.59% age 30-54 / 25.83% age
55+) — this file's embedded values matched every one of those exactly.
"""

import re

import openpyxl

import analysis_engine as ae
import menu_estimation as mest
import sales_data_builder as sdb


# Reuses sales_data_builder's CATEGORY_FC_BENCHMARK -- the same table
# input_normalizer.py and menu_estimation.py already use -- instead of
# keeping a separate local copy. This module previously defined its own
# table here with materially different values for the same category keys
# (e.g. SALAD 0.2514 vs 0.28, SAND/WRAP 0.1938 vs 0.30), which meant an
# identical raw item could get a different estimated Theoretical Cost
# purely depending on which file schema it arrived through (Menu
# Intelligence workbook vs. raw sales export vs. raw item list). Reusing
# sdb's table -- the one whose accuracy is documented and validated in
# the project README (52/54 exact match against the reference workbook)
# -- makes Theoretical Cost estimation consistent across every ingestion
# path. Replace sdb.CATEGORY_FC_BENCHMARK (in one place) if reusing this
# on a different restaurant's cost structure.
CATEGORY_FC_BENCHMARK = sdb.CATEGORY_FC_BENCHMARK
DEFAULT_FC_BENCHMARK = sdb.DEFAULT_FC_BENCHMARK

SUMMARY_ROW_MARKERS = ('total', 'portfolio', 'average', 'summary')

# canonical_name -> tuple of acceptable header spellings (case-insensitive,
# whitespace-insensitive). Add more variants here as new files show up.
COLUMN_ALIASES = {
    'name': ('menu item', 'current menu item', 'item', 'item name'),
    'category': ('category', 'current category', 'department'),
    'description': ('description', 'current description', 'menu description'),
    'ingredients': ('ingredients', 'current ingredients'),
    'annual_qty': ('total qty sold', 'total qty sold (annual)', 'total quantity sold', 'qty sold (annual)'),
    'popularity_percentile_given': ('popularity percentile (%)', 'popularity percentile', 'popularity percentile (0-100)'),
    'price': ('avg menu price ($)', 'avg menu price', 'current price ($)', 'price'),
    'theoretical_cost': ('theoretical cost ($)', 'theoretical cost'),
    'ingredient_cost': ('ingredient cost ($)', 'ingredient cost'),
    'prep_cost': ('prep cost ($)', 'prep cost'),
    'description_source': ('description source',),
    'ingredients_source': ('ingredients source',),
    'ingredient_cost_source': ('ingredient cost source',),
    'prep_cost_source': ('prep cost source',),
    'theoretical_cost_source': ('theoretical cost source', 'tc source'),
    'estimation_method': ('estimation method',),
    'estimation_source': ('estimation source',),
    'confidence_score': ('confidence score', 'confidence score (%)'),
    'zcta': ('zcta', 'zip', 'zip code'),
    'residents': ('living in area',),
    'daytime_workers': ('employed in area',),
    'worker_inflow': ('inflow jobs filled by outside workers',),
    'resident_outflow': ('outflow jobs filled by residents',),
    'stay_local': ('living and employed in area', 'interior jobs filled by residents'),
    'pct_income_high': ('inflow earning over 3333 pct',),
    'pct_income_low': ('inflow earning up to 1250 pct',),
    'pct_age_mid': ('inflow age 30 to 54 pct',),
    'pct_age_senior': ('inflow age 55 plus pct',),
}


def _norm(s):
    return re.sub(r'[^a-z0-9]+', ' ', str(s).strip().lower()).strip()


def _build_column_map(headers):
    """headers: list of raw header strings (index-aligned to columns).
    Returns {canonical_name: column_index} for whichever aliases matched."""
    norm_headers = [_norm(h) if h is not None else '' for h in headers]
    colmap = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        norm_aliases = [_norm(a) for a in aliases]
        for idx, nh in enumerate(norm_headers):
            if nh in norm_aliases:
                colmap[canonical] = idx
                break
    return colmap


def _find_header_row(ws, max_scan=10):
    for r in range(1, max_scan + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        norm_vals = {_norm(v) for v in vals if v}
        if norm_vals & {_norm(a) for a in COLUMN_ALIASES['name']}:
            return r, vals
    raise ValueError(
        "Could not find a header row (looked for a 'Menu Item' / 'Current "
        "Menu Item' column in the first 10 rows)."
    )


def _is_summary_row(name):
    if not name:
        return True
    n = str(name).strip().lower()
    return any(marker in n for marker in SUMMARY_ROW_MARKERS)


def resolve_theoretical_cost(price, given_tc, ingredient_cost, prep_cost, category,
                             name='', description='', ingredients=''):
    """Returns (theoretical_cost, source_label) per the priority order
    described in the module docstring."""
    if given_tc not in (None, '', 0) and price:
        try:
            given_tc = float(given_tc)
            fc_pct = given_tc / price
            if 0.05 <= fc_pct <= 0.70:
                return given_tc, 'given'
        except (TypeError, ValueError):
            pass

    if ingredient_cost not in (None, '') and prep_cost not in (None, ''):
        try:
            return float(ingredient_cost) + float(prep_cost), 'ingredient + prep cost'
        except (TypeError, ValueError):
            pass

    if name or description or ingredients:
        profile = mest.estimate_menu_item_profile(
            name=name, category=category, description=description, ingredients=ingredients,
            price=price, theoretical_cost=given_tc, ingredient_cost=ingredient_cost,
            prep_cost=prep_cost)
        return profile['theoretical_cost'], profile['theoretical_cost_source']

    benchmark = CATEGORY_FC_BENCHMARK.get(category)
    if benchmark is not None and price:
        return price * benchmark, f'category benchmark FC% ({benchmark*100:.1f}%)'

    if price:
        return price * DEFAULT_FC_BENCHMARK, f'default FC% ({DEFAULT_FC_BENCHMARK*100:.0f}%, no category benchmark available)'

    return 0.0, 'unresolved (no price to base an estimate on)'


def _cell(row_vals, colmap, canonical, default=None):
    idx = colmap.get(canonical)
    if idx is None or idx >= len(row_vals):
        return default
    v = row_vals[idx]
    return default if v is None else v


def extract_market_intel(rows_vals, colmap):
    """rows_vals: list of raw row-value lists. Looks for the first row with
    all the embedded commuter-flow columns present and returns a
    MarketIntel. Returns None if this schema's market columns aren't
    present at all (caller should fall back to manual entry / API fetch)."""
    needed = ('residents', 'daytime_workers', 'worker_inflow', 'resident_outflow',
              'stay_local', 'pct_income_high', 'pct_income_low', 'pct_age_mid', 'pct_age_senior')
    if not all(k in colmap for k in needed):
        return None

    for row in rows_vals:
        vals = {k: _cell(row, colmap, k) for k in needed}
        if all(v is not None for v in vals.values()):
            return ae.MarketIntel(
                residents=float(vals['residents']),
                daytime_workers=float(vals['daytime_workers']),
                worker_inflow=float(vals['worker_inflow']),
                resident_outflow=float(vals['resident_outflow']),
                stay_local=float(vals['stay_local']),
                pct_income_high=float(vals['pct_income_high']) * 100,
                pct_income_low=float(vals['pct_income_low']) * 100,
                pct_age_mid=float(vals['pct_age_mid']) * 100,
                pct_age_senior=float(vals['pct_age_senior']) * 100,
                pct_office_jobs=0.0,
            )
    return None


def load_menu_intelligence_workbook(file):
    """file: path or file-like object (an .xlsx). Returns:
      items_raw: list of dicts with keys name/category/ingredients/
                 annual_qty/price/theoretical_cost/tc_source
      market: MarketIntel or None (None => caller must supply one)
      warnings: list of str
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx, headers = _find_header_row(ws)
    colmap = _build_column_map(headers)

    warnings = []
    if 'name' not in colmap:
        raise ValueError("No recognizable 'Menu Item' column found — check the file's headers.")
    if 'category' not in colmap:
        warnings.append("No Category column found — every item will be treated as category 'UNKNOWN'.")
    if 'ingredients' not in colmap:
        warnings.append(
            "No Ingredients column in this file — duplicate and unique-ingredient "
            "detection will run on item name + category only (lower fidelity)."
        )
    if 'annual_qty' not in colmap:
        raise ValueError("No recognizable 'Total Qty Sold' column found — this is required.")
    if 'price' not in colmap:
        raise ValueError("No recognizable 'Avg Menu Price' / 'Price' column found — this is required.")

    all_rows_vals = []
    items_raw = []
    tc_estimated_count = 0
    for r in range(header_row_idx + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        name = _cell(vals, colmap, 'name')
        if name is None or _is_summary_row(name):
            continue
        all_rows_vals.append(vals)

        category = _cell(vals, colmap, 'category', default='UNKNOWN')
        price = _cell(vals, colmap, 'price')
        annual_qty = _cell(vals, colmap, 'annual_qty')
        ingredients = _cell(vals, colmap, 'ingredients', default='')

        if price is None or annual_qty is None:
            warnings.append(f"Skipped '{name}' — missing price or quantity sold.")
            continue

        given_tc = _cell(vals, colmap, 'theoretical_cost')
        ingredient_cost = _cell(vals, colmap, 'ingredient_cost')
        prep_cost = _cell(vals, colmap, 'prep_cost')
        tc, tc_source = resolve_theoretical_cost(
            float(price), given_tc, ingredient_cost, prep_cost, category)
        if tc_source != 'given':
            tc_estimated_count += 1

        items_raw.append({
            'name': str(name).strip(),
            'category': str(category).strip(),
            'ingredients': str(ingredients).strip() if ingredients else '',
            'annual_qty': float(annual_qty),
            'price': float(price),
            'theoretical_cost': tc,
            'tc_source': tc_source,
        })

    if tc_estimated_count:
        warnings.append(
            f"{tc_estimated_count} of {len(items_raw)} items had their Theoretical Cost "
            f"estimated (not taken directly from the file) — see each item's 'tc_source'."
        )

    market = extract_market_intel(all_rows_vals, colmap)
    if market is None:
        warnings.append(
            "No embedded market-intelligence columns found in this file — "
            "supply market data manually or via the Census/LODES fetch."
        )

    return items_raw, market, warnings