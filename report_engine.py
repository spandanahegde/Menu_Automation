#!/usr/bin/env python3
"""
build_report.py — Automated Grizz Grill Menu Refresh report generator.

Usage:
    python3 build_report.py <workbook.xlsx> [output.html] [--template report_template.html]

Reads the menu-refresh analysis workbook (same column layout as the source
"Menu Refresh" sheet), computes every derived figure the report needs, and
fills them into report_template.html to produce a client-ready HTML report.

The template's CSS, JS logic, and all non-data prose are never touched by
this script — only the tokens listed in TOKENS below are replaced. If the
template is edited to fix wording or styling, this script keeps working as
long as the token names stay the same.

ASSUMPTIONS (fail loudly if violated — see validate()):
  - The sheet has a header row containing the exact column names in
    COLUMN NAMES below (see REQUIRED_COLUMNS).
  - Each of the 54 (or however many) data rows has a unique "No." and a
    Recommendation of KEEP / REFRESH / REMOVE.
  - Item "No. 1" is used for the worked methodology examples in Sections
    03-06 (Popularity/Profitability/Commuter/Sales Forecast) — those
    sections still show that one item's numbers as a walkthrough and are
    NOT fully re-derived by this script beyond its Profitability Rank/tier
    (see LIMITATIONS at the bottom of this file).
"""

import sys
import re
import json
import html
import argparse
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl --break-system-packages")


# ----------------------------------------------------------------------
# Config: category display-name mapping and quintile tier labels.
# Extend CAT_MAP if the workbook introduces a new "Current Category" value.
# ----------------------------------------------------------------------
CAT_MAP = {
    'ENTREE': 'Entrée', 'SIDE / OTHER': 'Side', 'SALAD': 'Salad',
    'BREAKFAST': 'Breakfast', 'SAND / WRAP': 'Sandwich/Wrap',
    'APPETIZER': 'Appetizer', 'SOUP': 'Soup', 'BURGER': 'Burger',
    'BREAKFAST SIDE': 'Breakfast Side',
}

TIER_LABELS = {
    5: 'Highest Quintile (Top 20%)', 4: 'Above-Average Quintile',
    3: 'Mid-Range Quintile', 2: 'Below-Average Quintile',
    1: 'Lowest Quintile (Bottom 20%)',
}

REQUIRED_COLUMNS = [
    'No.', 'Current Menu Item', 'Current Category', 'Current Price ($)',
    'Theoretical Cost ($)', 'Current Margin %',
    'Profitability Rank (1-5, relative)', 'Commuter Score (0-100)',
    'Commuter Score Percentile', 'Demographic Multiplier',
    'Popularity Rank — Absolute', 'Popularity Percentile (0-100)',
    'Current Profit — 3 Mo ($)', 'Current Profit — 6 Mo ($)',
    'Current Profit — 9 Mo ($)', 'Current Profit — 12 Mo ($)',
    'Current Qty — Last 6 Mo', 'Total Forecast Qty — Next 6 Mo',
    'Qty Change — 6 Mo (Units)', 'Sales Change Signal',
    'Current Sales — Last 6 Mo ($)', 'Total Forecast Sales — Next 6 Mo ($)',
    'Sales Change — 6 Mo ($)', 'Recommendation (KEEP/REFRESH/REMOVE)',
    'Reason for Recommendation', 'New Ingredients Required',
    'Suggested Refreshed Item', 'Suggested Refreshed Description',
    'Procurement Impact', 'Operational Complexity',
    'Weekday Role', 'Weekend Role',
    # Validated Profitability/Sales-Forecast chain (analysis_engine.py) —
    # read and passed straight through to the popup modals so the JS never
    # has to re-derive Category Modifier / Market Multiplier itself. See
    # report_engine_patch_v2.md for why that re-derivation was a bug.
    'Category Modifier',
    'Future TC — 3 Mo ($)', 'Future TC — 6 Mo ($)', 'Future TC — 9 Mo ($)',
    'Future Margin — 3 Mo %', 'Future Margin — 6 Mo %', 'Future Margin — 9 Mo %',
    'Profitability Value — 3 Mo', 'Profitability Value — 6 Mo', 'Profitability Value — 9 Mo',
    'Baseline Growth Rate (%)', 'Market Multiplier', 'Adjusted Growth Rate (%)',
]


# ----------------------------------------------------------------------
# Extraction
# ----------------------------------------------------------------------
def load_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # find the header row: first row where the first cell is 'No.'
    header_row_idx = None
    for r in range(1, 10):
        if ws.cell(r, 1).value == 'No.':
            header_row_idx = r
            break
    if header_row_idx is None:
        sys.exit("Could not find header row (cell with 'No.' in column A "
                  "within the first 10 rows). Has the sheet layout changed?")

    headers = [ws.cell(header_row_idx, c).value for c in range(1, ws.max_column + 1)]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        sys.exit("Workbook is missing required column(s):\n  " + "\n  ".join(missing) +
                  "\nUpdate REQUIRED_COLUMNS / CAT_MAP in build_report.py if the "
                  "sheet's schema has intentionally changed.")

    rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if vals[0] is None:
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


# ----------------------------------------------------------------------
# Per-item derivation (mirrors the report's existing display conventions)
# ----------------------------------------------------------------------
def round_half_up(v):
    """Round-half-away-from-zero (matches Excel's ROUND() and JS Math.round()
    for positives) -- Python's builtin round() uses banker's rounding, which
    would silently disagree with the browser port on exact .5 values."""
    import math
    return int(math.floor(abs(v) + 0.5) * (1 if v >= 0 else -1))


def money0(v):
    return "{:,.0f}".format(round_half_up(v))


# ----------------------------------------------------------------------
# Section 02 — Market Movement & Customer Flow Intelligence.
# Previously 100% hardcoded to one restaurant's Memphis numbers; this
# builds the whole section body from whatever MarketIntel the pipeline
# actually resolved for this run (embedded in the file, fetched, or
# manually entered), so it's correct for any restaurant/ZCTA instead of
# always showing stale numbers from a different market.
# ----------------------------------------------------------------------
def build_market_section(market, restaurant_name='Grizz Grill', zcta=None):
    """market: an analysis_engine.MarketIntel, or None if no market data
    was available for this run (e.g. a bare finished-workbook upload with
    no embedded market columns) — in that case, returns a short, honest
    fallback instead of guessing or leaving stale numbers in place."""
    if market is None:
        return (
            '<p class="intro-p">Market and customer-flow data wasn\'t available '
            'for this run (no ZCTA or market data was provided), so this section '
            'is based on item-level performance only. Provide a ZIP code or '
            'market data to populate customer-flow intelligence here on the '
            'next run.</p>'
        )

    residents = market.residents
    daytime_workers = market.daytime_workers
    worker_inflow = market.worker_inflow
    resident_outflow = market.resident_outflow
    stay_local = market.stay_local
    pct_income_high = market.pct_income_high
    pct_income_low = market.pct_income_low
    pct_age_mid = market.pct_age_mid
    pct_age_senior = market.pct_age_senior

    pct_residents_leaving = (resident_outflow / residents * 100) if residents else 0.0
    worker_ratio = (daytime_workers / residents) if residents else 0.0
    net_inflow = worker_inflow - resident_outflow
    zcta_label = zcta or "this ZCTA"

    if residents > 0 and net_inflow > residents * 2:
        area_character = "daytime office workers far outnumbering local residents"
        net_interpretation = "confirming this is an extreme office-district location, not a residential neighborhood"
    elif net_inflow > 0:
        area_character = "a meaningful daytime commuter presence alongside its resident base"
        net_interpretation = "a net commuter-driven market on weekdays, though less lopsided than a pure office district"
    elif net_inflow == 0:
        area_character = "a roughly even balance of residents and daytime workers"
        net_interpretation = "a balanced market with no strong weekday commuter or residential skew"
    else:
        area_character = "a primarily residential population"
        net_interpretation = "a residential-driven market with more residents leaving for work than workers commuting in"

    esc_name = html.escape(restaurant_name)

    return f'''<p class="intro-p">Every recommendation in this report now factors in something most menu reviews miss: who is physically in the area, and when. {esc_name} sits in ZCTA {zcta_label} — a location shaped by {area_character}.</p>
<div class="kpi-grid" style="margin-top:26px;">
<div class="kpi-card"><div class="kpi-num total">{money0(residents)}</div><div class="kpi-label">Residents (Nighttime)</div></div>
<div class="kpi-card"><div class="kpi-num refresh">{money0(daytime_workers)}</div><div class="kpi-label">Daytime Workers Present</div></div>
<div class="kpi-card"><div class="kpi-num keep">{money0(worker_inflow)}</div><div class="kpi-label">Worker Inflow (Commuters In)</div></div>
<div class="kpi-card"><div class="kpi-num remove">{money0(resident_outflow)}</div><div class="kpi-label">Resident Outflow (Leave to Work)</div></div>
</div>
<h3 class="sub-head">What The Flow Looks Like</h3>
<div class="commuter-viz-wrap">
<svg aria-label="Commuter inflow and outflow diagram for ZCTA {zcta_label}" class="commuter-flow-svg" role="img" viewBox="0 0 680 300" xmlns="http://www.w3.org/2000/svg">
<title>Commuter inflow and outflow for ZCTA {zcta_label}</title>
<desc>A dark green arrow flowing into a circle representing {esc_name}'s ZCTA and a light green arrow flowing out of it, both at the same level, with a center pin marking the location</desc>
<circle cx="340" cy="150" r="78" fill="#97C459" fill-opacity="0.28" stroke="#639922" stroke-width="2"></circle>
<path d="M30,120 L250,120 L285,150 L250,180 L30,180 Z" fill="#27500A"></path>
<text x="145" y="150" text-anchor="middle" dominant-baseline="central" font-family="Bebas Neue, sans-serif" font-size="30" fill="#ffffff">{money0(worker_inflow)}</text>
<path d="M395,120 L615,120 L650,150 L615,180 L395,180 Z" fill="#97C459"></path>
<text x="530" y="150" text-anchor="middle" dominant-baseline="central" font-family="Bebas Neue, sans-serif" font-size="30" fill="#173404">{money0(resident_outflow)}</text>
<circle cx="340" cy="150" r="7" fill="#D14B2C" stroke="#ffffff" stroke-width="1.5"></circle>
<text x="340" y="177" text-anchor="middle" font-family="IBM Plex Mono, monospace" font-size="13" font-weight="700" fill="var(--smoke)">{zcta_label}</text>
<text x="145" y="215" text-anchor="middle" font-family="Inter, sans-serif" font-size="12" fill="var(--muted)">Workers commute in</text>
<text x="530" y="215" text-anchor="middle" font-family="Inter, sans-serif" font-size="12" fill="var(--muted)">Residents commute out</text>
<text x="340" y="252" text-anchor="middle" font-family="Inter, sans-serif" font-size="13" font-weight="700" fill="var(--smoke)">{money0(stay_local)} live &amp; work here</text>
</svg>
<div class="commuter-viz-legend">
<div class="cvl-item"><span class="cvl-dot" style="background:var(--keep-bg);"></span>Worker inflow ({money0(worker_inflow)}) is the dominant weekday signal every Weekday-role decision below is built on.</div>
<div class="cvl-item"><span class="cvl-dot" style="background:var(--gold);"></span>Resident outflow ({money0(resident_outflow)}) means the local resident base thins out on weekdays — {pct_residents_leaving:.1f}% of residents leave the area on weekdays, so weekend demand leans on the {money0(residents)} who stay.</div>
<div class="cvl-item"><span class="cvl-dot" style="background:var(--ember);"></span>Only {money0(stay_local)} people both live and work in {zcta_label} — the small "captive" segment present for both lunch and dinner occasions.</div>
</div>
</div>
<p class="flow-net">Net weekday inflow: <strong>+{money0(net_inflow)} people</strong> — {net_interpretation}.</p>
<h3 class="sub-head">The Trade Area, In Plain Numbers</h3>
<table class="data-table">
<thead><tr><th>Metric</th><th>Value</th><th>What It Means For {esc_name}</th></tr></thead>
<tbody>
<tr><td class="item-cell">Residents (nighttime population)</td><td class="mono">{money0(residents)}</td><td class="why-cell">The local base driving weekend and evening demand.</td></tr>
<tr><td class="item-cell">Daytime workers present</td><td class="mono">{money0(daytime_workers)}</td><td class="why-cell">{"Dominant weekday opportunity — " + f"{worker_ratio:.1f}x the resident base." if residents else "Weekday opportunity relative to the resident base."}</td></tr>
<tr><td class="item-cell">Workers commuting IN</td><td class="mono">{money0(worker_inflow)}</td><td class="why-cell">The core weekday lunch/breakfast target customer.</td></tr>
<tr><td class="item-cell">Residents commuting OUT</td><td class="mono">{money0(resident_outflow)}</td><td class="why-cell">{pct_residents_leaving:.1f}% of residents leave the area on weekdays — limits weekday resident traffic.</td></tr>
<tr><td class="item-cell">Live &amp; work locally</td><td class="mono">{money0(stay_local)}</td><td class="why-cell">Small but captive weekday + evening segment.</td></tr>
<tr><td class="item-cell">Inflow workers earning &gt;$3,333/mo</td><td class="mono">{pct_income_high:.1f}%</td><td class="why-cell">Spending power supports business-meal and premium pricing.</td></tr>
<tr><td class="item-cell">Inflow workers earning &lt;=$1,250/mo</td><td class="mono">{pct_income_low:.1f}%</td><td class="why-cell">A lower-wage segment — keep some low price-point options.</td></tr>
<tr><td class="item-cell">Inflow workers age 30-54</td><td class="mono">{pct_age_mid:.1f}%</td><td class="why-cell">Prime working-age professional crowd.</td></tr>
<tr><td class="item-cell">Inflow workers age 55+</td><td class="mono">{pct_age_senior:.1f}%</td><td class="why-cell">Secondary demographic segment to keep in mind for menu breadth.</td></tr>
</tbody>
</table>
<h3 class="sub-head">Weekday vs. Weekend — Two Different Restaurants</h3>
<div class="chart-row-2" style="margin-top:16px;">
<div class="cb healthy"><h5>Weekdays: Commuter-Driven</h5><p>Worker inflow ({money0(worker_inflow)}) {"dwarfs" if worker_inflow > residents * 2 else "compares to"} the resident base. Demand skews toward quick lunches, grab-and-go, and office-friendly items that a professional crowd can order fast.</p></div>
<div class="cb problem"><h5>Weekends: Resident-Driven</h5><p>With most commuters gone, local residents (just {money0(residents)}) become the primary segment. Demand shifts to family meals, shareables, and casual dining.</p></div>
</div>
<div class="callout"><strong>Strategic read:</strong> {"Weekday demand — breakfast, lunch, business meals, and takeaway — is the dominant revenue opportunity in this market. Weekend demand is comparatively smaller and residency-driven." if net_inflow > 0 else "Weekend and resident-driven demand is comparatively more significant here than in a typical commuter-heavy market — weigh family and casual-dining occasions accordingly."}</div>'''


def money2(v):
    return "{:,.2f}".format(v)


def sign_int(v):
    r = round_half_up(v)
    return ("+" if v >= 0 else "-") + str(abs(r))


# ----------------------------------------------------------------------
# Pricing Intelligence (Meal Occasion Pricing) — NEW, additive feature.
# Estimates occasion-specific (Lunch/Dinner/Weekday/Weekend) pricing and
# revenue from each item's real unit share, instead of one flat price.
#
# Like role_classifier.py, the daypart/weekend-role split percentages
# below are a judgment-call heuristic (not a validated formula like
# analysis_engine.py) — they use only fields already in the workbook
# (raw Category, Weekend Role) rather than requiring new upstream columns.
# The demand-multiplier brackets themselves ARE fixed or spec-exact (from
# the reference report). Only exposed for KEEP/REFRESH items — a REMOVE
# item has no forward price to optimize.
# ----------------------------------------------------------------------
CATEGORY_LUNCH_PCT = {
    'BREAKFAST': 90, 'BREAKFAST SIDE': 88, 'SAND / WRAP': 70, 'SALAD': 65,
    'SOUP': 60, 'SIDE / OTHER': 55, 'APPETIZER': 45, 'BURGER': 55,
    'ENTREE': 35,
}
DEFAULT_LUNCH_PCT = 50

WEEKEND_ROLE_PCT = {
    'Weekend Brunch': 45, 'Sports Event Special': 45, 'Family Meal': 42,
    'Premium Dinner': 40, 'Kids Favorite': 35, 'Casual Dining': 28.6,
}
DEFAULT_WEEKEND_PCT = 28.6  # 2/7 fair-share baseline


def lunch_multiplier(share):
    if share > 70:
        return 0.95
    if share >= 50:
        return 0.98
    if share >= 30:
        return 1.00
    return 1.03


def dinner_multiplier(share):
    if share > 70:
        return 1.05
    if share >= 50:
        return 1.03
    if share >= 30:
        return 1.00
    return 0.98


def weekday_multiplier(share):
    if share > 70:
        return 0.97
    if share >= 50:
        return 1.00
    return 1.03


def weekend_multiplier(share):
    if share > 70:
        return 1.05
    if share >= 50:
        return 1.03
    return 1.00


def compute_occasion_pricing(annual_qty, base_price, category_raw, weekend_role, est_total_units):
    """Returns a dict of every field the Pricing Intelligence popup needs,
    for all 4 occasions, computed once here rather than re-derived in JS."""
    lunch_pct = CATEGORY_LUNCH_PCT.get(category_raw, DEFAULT_LUNCH_PCT)
    dinner_pct = 100 - lunch_pct
    weekend_pct = WEEKEND_ROLE_PCT.get(weekend_role, DEFAULT_WEEKEND_PCT)
    weekday_pct = 100 - weekend_pct

    def occasion(share_pct, mult_fn):
        qty = annual_qty * share_pct / 100.0
        mult = mult_fn(share_pct)
        price = base_price * mult
        est_units = est_total_units * share_pct / 100.0
        est_rev = est_units * price
        return {
            'qty': round_half_up(qty), 'share': round(share_pct, 2), 'mult': round(mult, 4),
            'price': round(price, 2), 'est_units': round_half_up(est_units), 'est_rev': round(est_rev, 2),
        }

    return {
        'tq': round_half_up(annual_qty), 'etu': round_half_up(est_total_units), 'bp': round(base_price, 2),
        'lunch': occasion(lunch_pct, lunch_multiplier),
        'dinner': occasion(dinner_pct, dinner_multiplier),
        'weekday': occasion(weekday_pct, weekday_multiplier),
        'weekend': occasion(weekend_pct, weekend_multiplier),
    }


# ----------------------------------------------------------------------
# Section 06 Sales Forecast — pricing-driven recommendation columns
# (Lunch/Dinner, Weekday/Weekend, Adjusted Price (Final)). This is a
# DIFFERENT, simpler rule set than Pricing Intelligence's exact
# multipliers above — per spec, Section 06 uses ± percentage bands, not
# fixed multipliers, and picks a single best-selling-period recommendation
# per pair rather than showing all 4 occasions independently. Reuses the
# same Lunch/Dinner/Weekday/Weekend shares already computed in `pricing`
# (compute_occasion_pricing) so the two features never disagree on share.
# ----------------------------------------------------------------------
# ----------------------------------------------------------------------
# Section 06 Sales Forecast — Lunch/Dinner and Weekday/Weekend columns.
# Reads directly from the SAME `pricing` dict Pricing Intelligence uses
# (compute_occasion_pricing) — no separate rule set, so the two features
# can never show conflicting numbers.
# ----------------------------------------------------------------------
def build_item(r):
    no = r['No.']
    name = r['Current Menu Item']
    cat_raw = r['Current Category']
    cat = CAT_MAP.get(cat_raw, cat_raw.title())
    price = r['Current Price ($)']
    rec = r['Recommendation (KEEP/REFRESH/REMOVE)']
    s = rec.strip().lower()
    if s not in ('keep', 'refresh', 'remove'):
        sys.exit(f"Item #{no} ({name}) has an unrecognized Recommendation "
                  f"value: {rec!r}. Expected KEEP / REFRESH / REMOVE.")
    reason = r['Reason for Recommendation']

    wd = r['Weekday Role'] or ''
    we = r['Weekend Role'] or ''
    if '—' not in wd or '—' not in we:
        sys.exit(f"Item #{no} ({name}): Weekday/Weekend Role isn't in the "
                  f"expected 'Name — Description' format.")
    wdR, wdD = [x.strip() for x in wd.split('—', 1)]
    weR, weD = [x.strip() for x in we.split('—', 1)]

    pn = refresh_name = new_desc = new_ingredients = procurement = complexity = ""
    if s == 'refresh':
        refresh_name = r['Suggested Refreshed Item'] or ''
        srd = r['Suggested Refreshed Description'] or ''
        m = re.search(r'New price:\s*\$([\d,]+\.\d{2})', srd)
        if not m:
            sys.exit(f"Item #{no} ({name}) is REFRESH but no 'New price: $X.XX' "
                      f"found in Suggested Refreshed Description.")
        pn = m.group(1).replace(',', '')
        idx = srd.find('Reused from current recipe:')
        if idx == -1:
            idx = srd.find('Profitability Rank')
        desc_part = srd[:idx].strip() if idx != -1 else srd.strip()
        new_ingredients = r['New Ingredients Required'] or ''
        new_desc = desc_part + (f" New ingredients: {new_ingredients}." if new_ingredients else "")
        procurement = r['Procurement Impact'] or ''
        complexity = r['Operational Complexity'] or ''

    # Pricing Intelligence: only meaningful for items staying on the menu.
    # annual_qty and est_total_units are derived from fields already read
    # (Current Qty — Last 6 Mo is exactly half the annual total per
    # analysis_engine.py's compute_sales_forecast; est_total_units doubles
    # the item's own validated 6-mo forecast to approximate a forward
    # annual figure — no new upstream data required).
    raw_price = float(price)
    pricing = None
    if s in ('keep', 'refresh'):
        annual_qty = float(r['Current Qty — Last 6 Mo']) * 2
        est_total_units = float(r['Total Forecast Qty — Next 6 Mo']) * 2
        base_price_for_fc = float(pn) if s == 'refresh' else raw_price
        pricing = compute_occasion_pricing(annual_qty, base_price_for_fc, cat_raw, weR, est_total_units)

    return {
        'no': no, 'id': f'item-{no}', 'n': name, 'cat': cat,
        'p': money2(price), 'pn': pn, 's': s, 'why': reason,
        'pr': r['Popularity Rank — Absolute'], 'pp': r['Popularity Percentile (0-100)'],
        'mg': r['Current Margin %'], 'sp': money2(price), 'tc': money2(r['Theoretical Cost ($)']),
        'fr': r['Profitability Rank (1-5, relative)'], 'cs': r['Commuter Score (0-100)'],
        'dm': r['Demographic Multiplier'], 'mp': money2(r['Current Profit — 12 Mo ($)'] / 12),
        'raw_price': raw_price,
        'pricing': pricing,
        'p3': money0(r['Current Profit — 3 Mo ($)']), 'p6': money0(r['Current Profit — 6 Mo ($)']),
        'p9': money0(r['Current Profit — 9 Mo ($)']), 'p12': money0(r['Current Profit — 12 Mo ($)']),
        'uc': money0(r['Current Qty — Last 6 Mo']), 'uf': money0(r['Total Forecast Qty — Next 6 Mo']),
        'uch': sign_int(r['Qty Change — 6 Mo (Units)']), 'us': r['Sales Change Signal'],
        'sc': money0(r['Current Sales — Last 6 Mo ($)']), 'sf': money0(r['Total Forecast Sales — Next 6 Mo ($)']),
        'sch': str(round_half_up(r['Sales Change — 6 Mo ($)'])), 'ss': r['Sales Change Signal'],
        'wdR': wdR, 'wdD': wdD, 'weR': weR, 'weD': weD,
        'refresh_name': refresh_name, 'new_desc': new_desc,
        'new_ingredients': new_ingredients, 'procurement': procurement, 'complexity': complexity,
        'raw_sc': r['Current Sales — Last 6 Mo ($)'], 'raw_sf': r['Total Forecast Sales — Next 6 Mo ($)'],
        # Validated Profitability/Sales-Forecast chain, passed straight
        # through from analysis_engine.py — never recomputed here or in JS.
        'cm': r['Category Modifier'],
        'ftc3': r['Future TC — 3 Mo ($)'], 'ftc6': r['Future TC — 6 Mo ($)'], 'ftc9': r['Future TC — 9 Mo ($)'],
        'fm3': r['Future Margin — 3 Mo %'], 'fm6': r['Future Margin — 6 Mo %'], 'fm9': r['Future Margin — 9 Mo %'],
        'pv3': r['Profitability Value — 3 Mo'], 'pv6': r['Profitability Value — 6 Mo'],
        'pv9': r['Profitability Value — 9 Mo'],
        'bgr': r['Baseline Growth Rate (%)'], 'mm': r['Market Multiplier'],
        'agr': r['Adjusted Growth Rate (%)'],
    }


# ----------------------------------------------------------------------
# HTML fragment builders (accordions + sales-bridge table)
# ----------------------------------------------------------------------
def esc(s):
    return html.escape(str(s), quote=False)


def trend_pill(signal):
    return {'Increasing': 'trend-up', 'Decreasing': 'trend-down'}.get(signal, 'trend-flat')


def acc_metrics_block(it):
    tp = trend_pill(it['us'])
    rec_pill = {'keep': 'trend-up', 'refresh': 'trend-flat', 'remove': 'trend-down'}[it['s']]
    
    pricing_chip = ''
    if it.get('pricing'):
        base_p = it['pricing']['bp']
        pricing_chip = (
            f'<div class="acc-metric concept-trigger" data-concept="pricing" '
            f'data-item="{it["id"]}" data-period="lunch">'
            f'<span class="acc-metric-label">Pricing Intelligence &#9432;</span>'
            f'<span class="acc-metric-val">${base_p:.2f}</span>'
            f'</div>'
        )

    return f'''<div class="acc-metrics">
<div class="acc-metric concept-trigger" data-concept="popularity" data-item="{it['id']}"><span class="acc-metric-label">Popularity &#9432;</span><span class="acc-metric-val">{it['pr']}/54</span></div>

<div class="acc-metric concept-trigger" data-concept="profitability" data-item="{it['id']}"><span class="acc-metric-label">Profitability &#9432;</span><span class="acc-metric-val">{it['fr']}/5 &middot; {it['mg']:.2f}% margin</span></div>

<div class="acc-metric concept-trigger" data-concept="commuter" data-item="{it['id']}"><span class="acc-metric-label">Commuter Fit &#9432;</span><span class="acc-metric-val">{it['cs']:.1f}/100</span></div>

<div class="acc-metric concept-trigger" data-concept="salesforecast" data-item="{it['id']}"><span class="acc-metric-label">Sales Forecast &#9432;</span><span class="acc-metric-val">{it['uc']} &rarr; {it['uf']} <span class="trend-pill {tp}">{it['us']}</span></span></div>

{pricing_chip}

<div class="acc-metric concept-trigger" data-concept="recommendation" data-item="{it['id']}"><span class="acc-metric-label">Recommendation &#9432;</span><span class="acc-metric-val"><span class="trend-pill {rec_pill}">{it['s'].upper()}</span></span></div>

</div>'''

def cat_line(it):
    return (f'<p style="font-size:13px;color:var(--muted);margin:10px 0 6px;">'
            f'<strong>Category:</strong> {esc(it["cat"])} &nbsp;&middot;&nbsp; '
            f'<strong>Weekday:</strong> {esc(it["wdR"])} &mdash; {esc(it["wdD"])} &nbsp;&middot;&nbsp; '
            f'<strong>Weekend:</strong> {esc(it["weR"])} &mdash; {esc(it["weD"])}</p>')


def keep_block(it):
    return f'''<details class="acc acc-keep" id="{it['id']}">
<summary><div class="acc-left"><span class="acc-name">{esc(it['n'])}</span></div><div class="acc-right"><span class="acc-price-tag">${it['p']}</span><span class="acc-chevron">&#9654;</span></div></summary>
<div class="acc-body">
{acc_metrics_block(it)}
{cat_line(it)}
<div class="acc-why"><strong>Why we're suggesting this:</strong> {esc(it['why'])}</div>
</div>
</details>'''


def refresh_block(it):
    return f'''<details class="acc acc-refresh" id="{it['id']}">
<summary><div class="acc-left"><span class="acc-name">{esc(it['n'])}</span></div><div class="acc-right"><span class="acc-price-tag price-change">${it['p']} <span class="acc-price-arrow">&rarr;</span> ${it['pn']}</span><span class="acc-chevron">&#9654;</span></div></summary>
<div class="acc-body">
<p style="font-size:12.5px;color:var(--gold);font-weight:700;margin:10px 0 0;">becomes: {esc(it['refresh_name'])}</p>
{acc_metrics_block(it)}
{cat_line(it)}
<p style="font-size:13.5px;margin:10px 0;"><strong>New description:</strong> {esc(it['new_desc'])}</p>
<p style="font-size:12.5px;margin:8px 0 4px;"><strong>New ingredients:</strong> {esc(it['new_ingredients'])}</p>
<p style="font-size:12.5px;margin:4px 0 10px;"><strong>Procurement:</strong> {esc(it['procurement'])} &nbsp;&middot;&nbsp; <strong>Complexity:</strong> {esc(it['complexity'])}</p>
<div class="acc-why"><strong>Why we're suggesting this:</strong> {esc(it['why'])}</div>
</div>
</details>'''


def remove_block(it):
    return f'''<details class="acc acc-remove" id="{it['id']}">
<summary><div class="acc-left"><span class="acc-name">{esc(it['n'])}</span></div><div class="acc-right"><span class="acc-price-tag">${it['p']}</span><span class="acc-chevron">&#9654;</span></div></summary>
<div class="acc-body">
{acc_metrics_block(it)}
{cat_line(it)}
<div class="acc-why"><strong>Why we're suggesting this:</strong> {esc(it['why'])}</div>
</div>
</details>'''


def bridge_row(it):
    tp = trend_pill(it['us'])
    return f'''<tr>
<td><strong>{esc(it['n'])}</strong><br/><span style="font-size:11px;color:var(--muted);">{esc(it['cat'])}</span></td>
<td>{it['uc']}</td>
<td>${it['sc']}</td>
<td>{it['uf']}</td>
<td>${it['sf']}</td>
<td>{it['uch']}</td>
<td><span class="trend-pill {tp}">{it['us']}</span></td>
</tr>'''


def pricing_snapshot_row(it):
    """Table row of Lunch/Dinner + Weekday/Weekend recommendations for one
    item. Each recommendation cell and its matching New Price cell list
    occasions in the SAME order, one per line — one line (Both) or two
    lines (single winner) — so line N of the recommendation always lines
    up with line N of the price next to it. No separate price-band/
    engagement/adjustment columns; the calculation happens in
    compute_occasion_pricing() and only the final result is shown here.
    Returns '' for REMOVE items — no row rendered for them at all."""
    if it['pricing'] is None:
        return ''
    p = it['pricing']
    l, d, wd, we = p['lunch'], p['dinner'], p['weekday'], p['weekend']

    def pair_cells(a, b, label_a, label_b):
        # Near-balanced split (45-55%) — neither side meaningfully
        # dominates, so both occasions are recommended, each with its
        # own share and price, one below the other in matching order.
        if 45 <= a['share'] <= 55:
            rec = [f'{label_a} — {a["share"]:.0f}%', f'{label_b} — {b["share"]:.0f}%']
            price = [f'${money2(a["price"])}', f'${money2(b["price"])}']
        else:
            winner, w_label = (a, label_a) if a['share'] > b['share'] else (b, label_b)
            rec = [f'{w_label} — {winner["share"]:.0f}%']
            price = [f'${money2(winner["price"])}']
        return '<br/>'.join(rec), '<br/>'.join(price)

    ld_rec, ld_price = pair_cells(l, d, 'Lunch', 'Dinner')
    we_rec, we_price = pair_cells(wd, we, 'Weekday', 'Weekend')

    return f'''<tr>
<td><strong>{esc(it['n'])}</strong></td>
<td>${money2(p['bp'])}</td>
<td>{ld_rec}</td>
<td>{ld_price}</td>
<td>{we_rec}</td>
<td>{we_price}</td>
</tr>'''


# ----------------------------------------------------------------------
# Importable entry point — used by both the CLI (below) and the Streamlit
# app. Raises ReportBuildError with a human-readable message on bad input
# instead of calling sys.exit(), so a caller (e.g. Streamlit) can catch it
# and show it in the UI rather than crashing the process.
# ----------------------------------------------------------------------
class ReportBuildError(Exception):
    pass


def generate_report(workbook, template_path='report_template.html', market=None,
                     restaurant_name='Grizz Grill', zcta=None):
    """
    workbook: a path (str/Path) OR a file-like object (e.g. a Streamlit
              UploadedFile, or an open binary file handle) accepted by
              openpyxl.load_workbook.
    template_path: path to report_template.html.
    market: an analysis_engine.MarketIntel for Section 02 (Market Movement
            & Customer Flow Intelligence). If None, that section falls
            back to an honest "no market data available" note instead of
            showing stale numbers from a different run.
    restaurant_name / zcta: used to label Section 02. Both optional.

    Returns a dict: {
        'html': the finished report as a string,
        'counts': {'keep': int, 'refresh': int, 'remove': int, 'total': int},
        'warnings': [str, ...],   # non-fatal notes (e.g. item No.1 changed)
    }
    Raises ReportBuildError on any problem with the input data.
    """
    warnings = []
    try:
        rows = load_rows(workbook)
    except SystemExit as e:
        raise ReportBuildError(str(e))

    try:
        items = sorted((build_item(r) for r in rows), key=lambda i: i['no'])
    except SystemExit as e:
        raise ReportBuildError(str(e))
    n_total = len(items)

    keep = [i for i in items if i['s'] == 'keep']
    refresh = [i for i in items if i['s'] == 'refresh']
    remove = [i for i in items if i['s'] == 'remove']

    # --- validation ---
    counts = Counter(i['s'] for i in items)
    assert counts['keep'] + counts['refresh'] + counts['remove'] == n_total
    if n_total == 0:
        raise ReportBuildError("No data rows found — check the workbook.")

    # --- aggregates for summary/chart/bottom-line text ---
    def agg(group):
        return sum(i['raw_sc'] for i in group), sum(i['raw_sf'] for i in group)

    keep_sc, keep_sf = agg(keep)
    refresh_sc, refresh_sf = agg(refresh)
    remove_sc, remove_sf = agg(remove)

    pct_keep = round(len(keep) / n_total * 100)
    pct_refresh = round(len(refresh) / n_total * 100)
    pct_remove = round(len(remove) / n_total * 100)

    chart_lookfor = (
        f"{'Half' if pct_keep == 50 else str(pct_keep) + '%'} the menu "
        f"({len(keep)} of {n_total} items) already passes every signal in the "
        f"decision engine and needs no changes; the rest gets a targeted "
        f"refresh or comes off the menu."
    )
    chart_insight = (
        f"{pct_keep}% of the menu ({len(keep)} items) passes the 5-signal decision "
        f"engine cleanly — this model deliberately protects popular items from "
        f"removal unless multiple weak signals converge. {pct_refresh}% "
        f"({len(refresh)} items) get a targeted recipe refresh, mostly to fix a "
        f"below-median margin or below-median commuter fit, and {pct_remove}% "
        f"({len(remove)} items) fail enough signals at once to warrant removal."
    )
    remove_names = ', '.join(i['n'] for i in remove)
    remove_callout = (
        '<div class="callout">This item fails multiple converging signals at '
        'once — click for the full numbers.</div>'
        if len(remove) == 1 else
        '<div class="callout">These items fail multiple converging signals '
        '(or the ghost-item revenue floor) at once — click any item for the '
        'full numbers.</div>'
    )
    bottom_line = (
        "<p>Grizz Grill's menu is now shaped by a transparent, 5-signal decision "
        "engine — Popularity, Commuter Score, Current Profitability, Profitability "
        "Trend, and the 6-Month Sales Bridge — evaluated together for every item, "
        "with Profitability Rank scored relative to this menu's own margin "
        "distribution rather than fixed absolute bands, and with the reality of "
        "43,641 weekday commuters against just 6,462 residents baked into every "
        f"number. {len(keep)} of {n_total} items pass every signal cleanly and "
        f"should be left exactly as they are, projected to bring in roughly "
        f"${keep_sf:,.0f} in sales over the next 6 months (up from ${keep_sc:,.0f} "
        f"today). {len(refresh)} items show a real, specific concern — margin, "
        f"commuter fit, or brand duplication — now getting a targeted recipe and "
        f"price refresh projected to lift their 6-month sales from about "
        f"${refresh_sc:,.0f} to ${refresh_sf:,.0f}. "
        f"{len(remove)} item{'s' if len(remove) != 1 else ''} "
        f"({remove_names}) fail{'s' if len(remove) == 1 else ''} either the "
        f"$500 ghost-item revenue floor or two or more converging REMOVE "
        f"signals, representing about ${remove_sc:,.0f} in current 6-month "
        "sales combined. This model is deliberately conservative about removal "
        "— popularity alone protects an item, and only converging weak signals "
        "justify taking something off the menu.</p>"
    )

    # --- item-1 methodology example: now fully dynamic (Sections 04/06's
    # summary cards read every value from item1's actual validated fields,
    # no hardcoded name/numbers left in the template — see
    # report_engine_patch_v2.md) ---
    item1 = next((i for i in items if i['no'] == 1), items[0])
    item1_rank = item1['fr']
    item1_tier = TIER_LABELS.get(item1_rank, '—')

    # Profitability trend (Section 04's 9-Mo badge) — mirrors
    # profitabilityTrend() in the template's JS exactly (>1%/<1% band vs Stable).
    pv3, pv9 = item1['pv3'], item1['pv9']
    if pv9 > pv3 * 1.01:
        item1_trend_label, item1_trend_cls = 'Improving', 'pos'
    elif pv9 < pv3 * 0.99:
        item1_trend_label, item1_trend_cls = 'Declining', 'neg'
    else:
        item1_trend_label, item1_trend_cls = 'Stable', 'flat'

    # Sales Change Signal (Section 06's badges) — reuse the same
    # already-computed Sales Change Signal used everywhere else in the
    # report, so this can never disagree with it.
    item1_sales_change = item1['raw_sf'] - item1['raw_sc']
    item1_signal = item1['ss']
    item1_signal_cls = {'Increasing': 'pos', 'Decreasing': 'neg'}.get(item1_signal, 'flat')

    # --- render template ---
    with open(template_path, encoding='utf-8') as f:
        tpl = f.read()

    item_data = {i['id']: {k: (round(v, 4) if isinstance(v, float) else v)
                            for k, v in i.items()
                            if k in ('n', 's', 'p', 'pn', 'cat', 'wdR', 'wdD', 'weR', 'weD',
                                     'why', 'pr', 'pp', 'mg', 'sp', 'tc', 'fr', 'cs', 'dm',
                                     'mp', 'p3', 'p6', 'p9', 'p12', 'uc', 'uf', 'uch', 'us',
                                     'sc', 'sf', 'sch', 'ss',
                                     'cm', 'ftc3', 'ftc6', 'ftc9', 'fm3', 'fm6', 'fm9',
                                     'pv3', 'pv6', 'pv9', 'bgr', 'mm', 'agr', 'pricing')}
                 for i in items}

    tokens = {
        '{{KEEP_ACCORDION}}': '\n'.join(keep_block(i) for i in keep),
        '{{REFRESH_ACCORDION}}': '\n'.join(refresh_block(i) for i in refresh),
        '{{REMOVE_ACCORDION}}': '\n'.join(remove_block(i) for i in remove),
        '{{SALES_BRIDGE_ROWS}}': '\n'.join(bridge_row(i) for i in items),
        '{{PRICING_SNAPSHOT_ROWS}}': '\n'.join(
            row for row in (pricing_snapshot_row(i) for i in items) if row),
        '{{ITEM_DATA_JSON}}': json.dumps(item_data, ensure_ascii=False, separators=(',', ':')),
        '{{KEEP_COUNT}}': str(len(keep)),
        '{{REFRESH_COUNT}}': str(len(refresh)),
        '{{REMOVE_COUNT}}': str(len(remove)),
        '{{REMOVE_CALLOUT}}': remove_callout,
        '{{CHART_LOOKFOR_TEXT}}': chart_lookfor,
        '{{CHART_INSIGHT_TEXT}}': chart_insight,
        '{{BOTTOM_LINE_PARAGRAPH}}': bottom_line,
        '{{ITEM1_NAME}}': html.escape(item1['n']),
        '{{ITEM1_MARGIN_PCT}}': f"{item1['mg']:.2f}%",
        '{{ITEM1_PV_3MO}}': f"{item1['pv3']:.2f}",
        '{{ITEM1_FM_3MO}}': f"{item1['fm3']:.2f}%",
        '{{ITEM1_PV_6MO}}': f"{item1['pv6']:.2f}",
        '{{ITEM1_FM_6MO}}': f"{item1['fm6']:.2f}%",
        '{{ITEM1_PV_9MO}}': f"{item1['pv9']:.2f}",
        '{{ITEM1_TREND_LABEL}}': item1_trend_label,
        '{{ITEM1_TREND_BADGE_CLASS}}': item1_trend_cls,
        '{{ITEM1_SALES_6MO}}': f"${item1['sc']}",
        '{{ITEM1_UNITS_6MO}}': item1['uc'],
        '{{ITEM1_FORECAST_SALES_6MO}}': f"${item1['sf']}",
        '{{ITEM1_SALES_CHANGE}}': (('+' if item1_sales_change >= 0 else '-')
                                    + '$' + money0(abs(item1_sales_change))),
        '{{ITEM1_FORECAST_UNITS_6MO}}': item1['uf'],
        '{{ITEM1_SIGNAL}}': item1_signal,
        '{{ITEM1_SIGNAL_BADGE_CLASS}}': item1_signal_cls,
        '{{ITEM1_RANK}}': str(item1_rank),
        '{{ITEM1_TIER_LABEL}}': item1_tier,
        '{{RESTAURANT_NAME}}': html.escape(restaurant_name),
        '{{MARKET_SECTION_BODY}}': build_market_section(market, restaurant_name, zcta),
    }

    out = tpl
    for tok, val in tokens.items():
        if tok not in out:
            raise ReportBuildError(
                f"Template is missing expected token {tok} — has the "
                f"template been edited? Re-check report_template.html.")
        out = out.replace(tok, val)

    leftover = re.findall(r'\{\{[A-Z0-9_]+\}\}', out)
    if leftover:
        raise ReportBuildError(f"Unfilled template tokens remain: {set(leftover)}")

    # --- structural sanity checks before returning ---
    assert out.count('<details') == out.count('</details') == n_total
    assert out.count('<section') == out.count('</section>')
    json.loads(re.search(r'const ITEM_DATA = (\{.*?\});', out, re.S).group(1))

    return {
        'html': out,
        'counts': {'keep': len(keep), 'refresh': len(refresh),
                   'remove': len(remove), 'total': n_total},
        'warnings': warnings,
    }


# ----------------------------------------------------------------------
# CLI wrapper (unchanged usage: python3 report_engine.py workbook.xlsx out.html)
# ----------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('workbook', help='Path to the menu-refresh .xlsx file')
    ap.add_argument('output', nargs='?', default='Menu_Refresh_Report.html',
                     help='Output HTML path (default: Menu_Refresh_Report.html)')
    ap.add_argument('--template', default='report_template.html',
                     help='Path to the HTML template (default: report_template.html)')
    args = ap.parse_args()

    try:
        result = generate_report(args.workbook, template_path=args.template)
    except ReportBuildError as e:
        sys.exit(str(e))

    for w in result['warnings']:
        print(f"NOTE: {w}", file=sys.stderr)

    with open(args.output, 'w', encoding='utf-8') as f:
        f.write(result['html'])

    c = result['counts']
    print(f"Wrote {args.output}: {c['keep']} Keep / {c['refresh']} Refresh / "
          f"{c['remove']} Remove (of {c['total']} items).")


if __name__ == '__main__':
    main()


# ----------------------------------------------------------------------
# LIMITATIONS (read before trusting this for a very different menu)
# ----------------------------------------------------------------------
# 1. Sections 04 (Profitability) and 06 (Sales Forecast) summary cards are
#    now fully dynamic — every value comes from item No.1's actual
#    validated fields (Category Modifier, Future Margin/TC, Profitability
#    Value 3/6/9-Mo, Baseline/Adjusted Growth Rate, Market Multiplier — all
#    computed by analysis_engine.py, never re-derived here or in the
#    template's JS). Sections 03 (Popularity) and 05 (Commuter Score)
#    still show item No.1's numbers as literal static text in the
#    template — those weren't in scope for this pass. If they need the
#    same treatment, the pattern to follow is identical: add the needed
#    columns to REQUIRED_COLUMNS, read them in build_item(), add
#    {{ITEM1_...}} tokens here, and replace the corresponding hardcoded
#    numbers in report_template.html's Section 03/05 msect-kpi-grid.
# 2. The demographic/market section (ZCTA figures, resident/worker counts)
#    is treated as fixed restaurant context and is never touched. If the
#    trade-area data itself changes, that section needs a manual edit.
# 3. Category names not in CAT_MAP fall back to Python's .title() — add
#    an explicit mapping if a new raw category label shows up.
# 4. Every item's popup modal (not just item No.1's) reads its Category
#    Modifier / Profitability Value / Market Multiplier straight from the
#    workbook via ITEM_DATA_JSON — the template's JS no longer guesses
#    these from the item's name/category text. See report_engine_patch_v2.md
#    for the bug this fixed (guessed category modifiers could silently
#    disagree with the validated ones for any item outside a few
#    hardcoded ENTREE overrides).