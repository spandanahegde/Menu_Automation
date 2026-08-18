"""
workflow_orchestrator.py — the actual product: one function that runs the
complete Menu Refresh process from whatever input you hand it through to
the three outputs (refreshed menu decisions, workbook, report). This is
what "Start Menu Refresh" in the UI calls — the four separate ingestion
paths (finished workbook / raw item list / Menu Intelligence workbook /
raw POS sales export) are implementation detail this module auto-detects
and hides, not something the person running it should have to know about.

    run_menu_refresh(file, zcta=None, market_override=None,
                     progress_callback=None) -> WorkflowResult

Sequence (matches the intended pipeline):
  1. Detect what kind of file this is (four possible schemas).
  2. If it's raw sales data or a raw item list, get market intelligence
     (embedded in the file if present, otherwise fetch Census + LODES
     automatically using the ZCTA -- either extracted from the file or
     passed in -- falling back to market_override if fetching fails and
     one was supplied).
  3. Run role classification, duplicate detection, and the full Sections
     A-I engine.
  4. For every REMOVE item, generate a replacement brief (the gap left
     behind and what a new item should target) -- REMOVE withacout a
     replacement direction is an incomplete decision, not a finished one.
  5. Build the Menu Refresh workbook and the HTML report.

Every step reports its progress through progress_callback(message,
fraction) if one is given, so a caller (the Streamlit app) can show it
live instead of a silent black box.
"""

import io
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
import pandas as pd

import analysis_engine as ae
import role_classifier as rc
import duplicate_detector as dd
import menu_intelligence_ingest as mii
import sales_data_builder as sdb
import run_analysis as ra
import replacement_recommender as rr
import report_engine
import market_data as md
import input_normalizer as inorm


class WorkflowError(Exception):
    pass


@dataclass
class WorkflowResult:
    input_type: str
    items: list
    duplicate_candidates: list
    replacement_briefs: list
    market: object
    market_source: str
    workbook_bytes: object
    menu_intelligence_bytes: Optional[object]
    report_html: str
    counts: dict
    warnings: list = field(default_factory=list)


def _noop(message, fraction=None):
    pass


def _get_bytes(file):
    """Returns (raw_bytes, filename). Works for a path string, an open
    file handle, or a Streamlit UploadedFile."""
    if isinstance(file, (bytes, bytearray)):
        return bytes(file), getattr(file, 'name', 'upload')
    if isinstance(file, str):
        with open(file, 'rb') as f:
            return f.read(), file
    name = getattr(file, 'name', 'upload')
    file.seek(0)
    data = file.read()
    file.seek(0)
    return data, name


def _peek_headers(raw_bytes, filename, header_scan_rows=10):
    """Returns a set of normalized header strings found anywhere in the
    first `header_scan_rows` rows — used to fingerprint which schema this
    file is, without committing to one parser yet."""
    headers = set()
    if filename.lower().endswith('.csv'):
        text = raw_bytes.decode('utf-8', errors='replace')
        first_lines = text.splitlines()[:header_scan_rows]
        for line in first_lines:
            for cell in line.split(','):
                headers.add(cell.strip().strip('"'))
    else:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in range(1, min(header_scan_rows, ws.max_row) + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    headers.add(str(v).strip())
    return headers


def detect_input_type(file):
    """Returns one of: 'finished_workbook', 'menu_intelligence',
    'raw_item_list', 'sales_export'. Raises WorkflowError if nothing matches.

    The first three checks look for exact, unambiguous signature columns
    that only ever appear in one specific schema (a finished workbook's
    Recommendation column, a POS export's MICROS-style field names, a
    Menu Intelligence workbook's cost-breakdown columns) — those stay
    exact-match on purpose, so a file that happens to share a header name
    doesn't get misclassified into the wrong pipeline.

    The last check is deliberately broad: ANY file with a recognizable
    item name + price + quantity-sold column (via input_normalizer's
    alias table — 'Item Name'/'Selling Price'/'Quantity Sold' and many
    other real-world variants all match) is accepted as 'raw_item_list',
    even with nothing else. Category/Ingredients/Theoretical Cost get
    generated downstream if they're missing — see input_normalizer.py.
    """
    raw_bytes, filename = _get_bytes(file)
    headers = _peek_headers(raw_bytes, filename)

    if 'Recommendation (KEEP/REFRESH/REMOVE)' in headers and 'No.' in headers:
        return 'finished_workbook'

    if {'MENU_ITEM_NAME', 'FAMILY_GROUP_NAME', 'STORE_QTY_SOLD'}.issubset(headers):
        return 'sales_export'

    if 'Ingredient Cost ($)' in headers or 'Prep Cost ($)' in headers:
        return 'menu_intelligence'

    if inorm.looks_like_minimal_schema(headers):
        return 'raw_item_list'

    raise WorkflowError(
        "Couldn't recognize this file's format. At minimum, it needs a "
        "recognizable item-name column, a price column, and a "
        "quantity-sold column (many header spellings are recognized — "
        "see input_normalizer.CANONICAL_ALIASES for the current list, or "
        "add yours there). Headers found in this file: "
        f"{sorted(h for h in headers if h)}."
    )


def _get_market(zcta, progress_callback):
    """Automatic market-data fetch: Census ACS5 + LODES. Raises
    WorkflowError with a clear message on failure — caller decides
    whether to fall back to a manually-supplied market_override."""
    if not zcta:
        raise WorkflowError("No ZCTA available to fetch market data with.")
    progress_callback(f"Fetching Census demographics for ZCTA {zcta}…", 0.15)
    demo = md.fetch_census_demographics(zcta)
    progress_callback(f"Got resident population ({demo.get('residents_source','ACS')}). "
                       f"Fetching commuter flow data (LODES)…", 0.20)
    state = md.zcta_to_state(zcta)
    if not state:
        raise WorkflowError(f"Couldn't determine a state for ZCTA {zcta}.")
    flows = md.fetch_commuter_flows(zcta, state, progress_callback=progress_callback)
    return ae.MarketIntel(
        residents=demo['residents'], daytime_workers=flows['daytime_workers'],
        worker_inflow=flows['worker_inflow'], resident_outflow=flows['resident_outflow'],
        stay_local=flows['stay_local'], pct_income_high=flows['pct_income_high'],
        pct_income_low=flows['pct_income_low'], pct_age_mid=flows['pct_age_mid'],
        pct_age_senior=flows['pct_age_senior'], pct_office_jobs=0.0,
    ), flows.get('source', 'LODES') + ' + ' + demo.get('residents_source', 'ACS')


def _extract_zcta_from_sales_csv(file):
    raw_bytes, filename = _get_bytes(file)
    try:
        if filename.lower().endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(raw_bytes), usecols=lambda c: c == 'ZCTA')
        else:
            df = pd.read_csv(io.StringIO(raw_bytes.decode('utf-8', errors='replace')), usecols=lambda c: c == 'ZCTA')
        if 'ZCTA' in df.columns:
            vals = df['ZCTA'].dropna()
            if len(vals):
                return str(int(vals.mode().iloc[0]))
    except Exception:
        pass
    return None


def run_menu_refresh(file, zcta=None, market_override=None, template_path='report_template.html',
                      progress_callback=None, restaurant_name='Grizz Grill'):
    """The single entry point. See module docstring for the sequence.
    Returns a WorkflowResult. Raises WorkflowError on anything that
    stops the whole run (unrecognized file, no usable market data)."""
    progress = progress_callback or _noop

    progress("Reading the file…", 0.02)
    input_type = detect_input_type(file)
    progress(f"Recognized this as a {input_type.replace('_', ' ')}.", 0.05)

    warnings = []
    menu_intelligence_bytes = None
    zcta_used = zcta

    if input_type == 'finished_workbook':
        progress("This is already a finished analysis workbook — building the report directly.", 0.5)
        raw_bytes, _ = _get_bytes(file)

        market = market_override
        market_source = 'manually supplied' if market is not None else None
        if market is None and zcta:
            try:
                market, market_source = _get_market(zcta, progress)
            except WorkflowError as e:
                warnings.append(f"Live market-data fetch for ZCTA {zcta} failed ({e}) — "
                                 f"the report's market section will show item-level data only.")
        if market is None:
            market_source = '(none provided — finished workbooks have no embedded market data)'

        result = report_engine.generate_report(io.BytesIO(raw_bytes), template_path=template_path,
                                                 restaurant_name=restaurant_name, market=market,
                                                 zcta=zcta if market is not None else None)
        progress("Report generated. This path skips item-level replacement "
                 "briefs — those need the underlying item objects, which a "
                 "finished workbook doesn't expose the same way.", 0.95)
        return WorkflowResult(
            input_type=input_type, items=[], duplicate_candidates=[], replacement_briefs=[],
            market=market, market_source=market_source,
            workbook_bytes=io.BytesIO(raw_bytes), menu_intelligence_bytes=None,
            report_html=result['html'], counts=result['counts'], warnings=result['warnings'] + warnings,
        )

    if input_type == 'menu_intelligence':
        progress("Reading embedded financial + market data…", 0.10)
        raw_bytes, _ = _get_bytes(file)
        market = market_override
        market_source = 'manually supplied'
        if market is None:
            _, market_from_file, _ = mii.load_menu_intelligence_workbook(io.BytesIO(raw_bytes))
            market = market_from_file
            market_source = 'embedded in file'
        if market is None:
            if zcta:
                market, market_source = _get_market(zcta, progress)
            else:
                raise WorkflowError(
                    "This file has no embedded market data, no market_override was given, "
                    "and no ZCTA was provided to fetch one automatically."
                )
        progress("Running role classification, duplicate detection, and the full engine…", 0.45)
        items, pairs, market_used, ingest_warnings = ra.build_items_from_menu_intelligence(
            io.BytesIO(raw_bytes), market_override=market)
        warnings.extend(ingest_warnings)
        fin_items = None

    elif input_type == 'sales_export':
        progress("Building Menu Intelligence financial fields from the raw sales data…", 0.10)
        raw_bytes, _ = _get_bytes(file)
        market = market_override
        market_source = 'manually supplied'
        if market is None:
            file_zcta = zcta or _extract_zcta_from_sales_csv(io.BytesIO(raw_bytes))
            zcta_used = file_zcta or zcta_used
            if file_zcta:
                try:
                    market, market_source = _get_market(file_zcta, progress)
                except WorkflowError as e:
                    if market_override is not None:
                        market, market_source = market_override, 'manually supplied (auto-fetch failed)'
                    else:
                        raise WorkflowError(
                            f"Automatic market-data fetch failed for ZCTA {file_zcta}: {e}. "
                            f"Supply market_override to proceed anyway."
                        )
            else:
                raise WorkflowError(
                    "No ZCTA found in the file and none was provided — can't fetch market data."
                )
        progress("Running role classification, duplicate detection, and the full engine…", 0.45)
        items, pairs, fin_items, ingest_warnings = ra.build_items_from_sales_data(
            io.BytesIO(raw_bytes), market_override=market)
        warnings.extend(ingest_warnings)
        mi_wb = ra.build_menu_intelligence_workbook_bytes(fin_items, market)
        menu_intelligence_bytes = mi_wb

    elif input_type == 'raw_item_list':
        progress("Reading and normalizing the item list…", 0.10)
        raw_bytes, filename = _get_bytes(file)
        df = (pd.read_excel(io.BytesIO(raw_bytes)) if filename.lower().endswith('.xlsx')
              else pd.read_csv(io.StringIO(raw_bytes.decode('utf-8', errors='replace'))))
        try:
            norm_result = inorm.normalize_dataframe(df)
        except ValueError as e:
            raise WorkflowError(str(e))
        warnings.extend(norm_result.warnings)
        raw_rows = norm_result.rows
        if not raw_rows:
            raise WorkflowError("No usable rows after normalization — check the file's data.")

        market = market_override
        market_source = 'manually supplied'
        if market is None:
            effective_zcta = zcta
            if not effective_zcta and 'zcta' in norm_result.column_map:
                zcta_vals = df[norm_result.column_map['zcta']].dropna()
                if len(zcta_vals):
                    effective_zcta = str(int(zcta_vals.mode().iloc[0]))
            if effective_zcta:
                zcta_used = effective_zcta
                market, market_source = _get_market(effective_zcta, progress)
            else:
                raise WorkflowError(
                    "This file has no embedded market data and no ZCTA was provided to fetch one."
                )
        progress("Running role classification, duplicate detection, and the full engine…", 0.45)
        items, pairs = ra.build_items_from_raw(raw_rows, market)
        fin_items = None

    else:
        raise WorkflowError(f"Unhandled input type: {input_type}")

    progress("Generating replacement briefs for any REMOVE items…", 0.65)
    briefs = rr.build_replacement_briefs(items)

    progress("Building the Menu Refresh workbook (drafting REFRESH details)…", 0.75)
    llm_warnings = []
    workbook_bytes = ra.build_workbook_bytes(items, use_llm=True, llm_warnings=llm_warnings)
    if llm_warnings:
        n_refresh = sum(1 for it in items if it.recommendation == 'REFRESH')
        progress(f"{len(llm_warnings)} of {n_refresh} REFRESH item(s) used the deterministic "
                 f"content generator (no LLM available/valid).", 0.80)
        if len(llm_warnings) <= 3:
            warnings.extend(llm_warnings)
        else:
            reason = llm_warnings[0].split('(', 1)[-1].rstrip(').') if '(' in llm_warnings[0] else 'see workbook'
            warnings.append(
                f"{len(llm_warnings)} of {n_refresh} REFRESH item(s) used the deterministic "
                f"content generator instead of the LLM ({reason}) — every field is still "
                f"complete (name, description, ingredients, cost, procurement), just not "
                f"LLM-tailored. Set ANTHROPIC_API_KEY for genuinely bespoke creative content."
            )

    progress("Generating the executive HTML report…", 0.85)
    result = report_engine.generate_report(
        workbook_bytes, template_path=template_path, market=market,
        restaurant_name=restaurant_name, zcta=zcta_used)
    warnings.extend(result['warnings'])

    progress("Done.", 1.0)
    return WorkflowResult(
        input_type=input_type, items=items, duplicate_candidates=pairs,
        replacement_briefs=briefs, market=market, market_source=market_source,
        workbook_bytes=workbook_bytes, menu_intelligence_bytes=menu_intelligence_bytes,
        report_html=result['html'], counts=result['counts'], warnings=warnings,
    )